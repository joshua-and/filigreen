#!/usr/bin/env python3
"""
Export bank_for_daos_capital_tracking.xlsx to a single CSV in accounting format,
and merge in transaction history from Ethereum Mainnet and Base export CSVs.
Output: raw/filigreen_accounting_export.csv
"""

import csv
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("Install dependencies: pip install -r requirements.txt")

REPO_ROOT = Path(__file__).resolve().parent.parent
RAW = REPO_ROOT / "raw"
EXCEL_PATH = RAW / "bank_for_daos_capital_tracking.xlsx"
ETH_CSV = RAW / "transactions_export_1_0xeDC0F81bbD4776dD3194b19D5e45a5d22D77408E_1771846211502.csv"
BASE_CSV = RAW / "transactions_export_8453_0xeDC0F81bbD4776dD3194b19D5e45a5d22D77408E_1771846265783.csv"
OUTPUT_PATH = RAW / "filigreen_accounting_export.csv"

# Unified CSV columns (Section first, then section-specific; empty cells for N/A)
HEADERS = [
    "Section",
    "Member Name",
    "Member Type",
    "Total Capital",
    "Loan 1 Allocation %",
    "Loan 2 Allocation %",
    "Allocation Notes",
    "Month",
    "Loan 1 Interest",
    "Loan 1 Principal",
    "Loan 1 Fees/Expenses",
    "Loan 1 Balance",
    "Loan 2 Interest",
    "Loan 2 Principal",
    "Loan 2 Fees/Expenses",
    "Loan 2 Balance",
    "Loan",
    "Detail Field",
    "Detail Value",
    "Date",
    "Chain",
    "From Address",
    "To Address",
    "Amount",
    "Asset Type",
    "Asset Symbol",
    "Transaction Hash",
    "Contract Address",
    "Note",
]


def row_dict(section, **kwargs):
    out = {h: "" for h in HEADERS}
    out["Section"] = section
    for k, v in kwargs.items():
        if k in out and v is not None:
            out[k] = "" if v is None else str(v).strip()
    return out


def export_member_allocations(ws, rows):
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and "Member Name" in (str(c) for c in row if c is not None):
            header_row_idx = i
            break
    if header_row_idx is None:
        return
    row_1based = header_row_idx + 1
    headers = [
        str(ws.cell(row=row_1based, column=c).value or "").strip()
        for c in range(1, ws.max_column + 1)
    ]
    col_map = {
        "Member Name": "Member Name",
        "Member Type": "Member Type",
        "Total Capital": "Total Capital",
        "Loan 1 Allocation %": "Loan 1 Allocation %",
        "Loan 2 Allocation %": "Loan 2 Allocation %",
        "Notes": "Allocation Notes",
    }
    for r in range(header_row_idx + 2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue
        if any("TOTAL" in str(v or "") for v in row_vals):
            continue
        if str(row_vals[0] or "").strip().startswith(("Instructions", "1.", "2.", "3.")):
            continue
        d = row_dict("Member Allocations")
        for i, h in enumerate(headers):
            if h in col_map and i < len(row_vals) and row_vals[i] is not None:
                d[col_map[h]] = str(row_vals[i]).strip()
        rows.append([d[h] for h in HEADERS])


def export_loan_performance(ws, rows):
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and "Month" in (str(c) for c in row if c is not None):
            header_row_idx = i
            break
    if header_row_idx is None:
        return
    row_1based = header_row_idx + 1
    headers = [
        str(ws.cell(row=row_1based, column=c).value or "").strip()
        for c in range(1, ws.max_column + 1)
    ]
    col_map = {
        "Month": "Month",
        "Loan 1 Interest": "Loan 1 Interest",
        "Loan 1 Principal": "Loan 1 Principal",
        "Loan 1 Fees/Expenses": "Loan 1 Fees/Expenses",
        "Loan 1 Balance": "Loan 1 Balance",
        "Loan 2 Interest": "Loan 2 Interest",
        "Loan 2 Principal": "Loan 2 Principal",
        "Loan 2 Fees/Expenses": "Loan 2 Fees/Expenses",
        "Loan 2 Balance": "Loan 2 Balance",
    }
    for r in range(header_row_idx + 2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue
        if any("ANNUAL TOTAL" in str(v or "") for v in row_vals):
            continue
        if str(row_vals[0] or "").strip().startswith(("Instructions", "1.", "2.", "3.", "4.", "5.", "Data Sources", "- ")):
            continue
        d = row_dict("Loan Performance")
        for i, h in enumerate(headers):
            if h in col_map and i < len(row_vals) and row_vals[i] is not None:
                d[col_map[h]] = str(row_vals[i]).strip()
        rows.append([d[h] for h in HEADERS])


def export_loan_details(ws, rows):
    current_loan = None
    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row if c is not None]
        if not cells:
            continue
        first = str(cells[0]).strip()
        if first == "LOAN 1":
            current_loan = "1"
            continue
        if first == "LOAN 2":
            current_loan = "2"
            continue
        if first in ("Instructions", "Instructions:") or first.startswith("1."):
            continue
        if current_loan and first.endswith(":"):
            field = first.rstrip(":")
            value = str(cells[1]).strip() if len(cells) > 1 else ""
            d = row_dict("Loan Details", Loan=current_loan)
            d["Detail Field"], d["Detail Value"] = field, value
            rows.append([d[h] for h in HEADERS])
        elif current_loan and len(cells) >= 2 and first not in ("Instructions", "Instructions:"):
            field = first
            value = str(cells[1]).strip() if len(cells) > 1 else ""
            if field and "Loan" not in field and not field.startswith("1."):
                d = row_dict("Loan Details", Loan=current_loan)
                d["Detail Field"], d["Detail Value"] = field, value
                rows.append([d[h] for h in HEADERS])


def load_transaction_csv(path, chain_name):
    rows_out = []
    with open(path, newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for rec in r:
            date = (rec.get("Executed at") or rec.get("Created at") or "").split("T")[0]
            d = row_dict("Transactions", Date=date, Chain=chain_name)
            d["From Address"] = rec.get("From Address") or ""
            d["To Address"] = rec.get("To Address") or ""
            d["Amount"] = rec.get("Amount") or ""
            d["Asset Type"] = rec.get("Asset Type") or ""
            d["Asset Symbol"] = rec.get("Asset Symbol") or ""
            d["Transaction Hash"] = rec.get("Transaction Hash") or ""
            d["Contract Address"] = rec.get("Contract Address") or ""
            d["Note"] = rec.get("Note") or ""
            rows_out.append([d[h] for h in HEADERS])
    return rows_out


def main():
    RAW.mkdir(parents=True, exist_ok=True)
    if not EXCEL_PATH.exists():
        raise SystemExit(f"Excel file not found: {EXCEL_PATH}")

    all_rows = []
    wb = load_workbook(EXCEL_PATH, data_only=True)
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            if "Member" in name:
                export_member_allocations(ws, all_rows)
            elif "Performance" in name:
                export_loan_performance(ws, all_rows)
            elif "Details" in name or "Loan Details" in name:
                export_loan_details(ws, all_rows)
    finally:
        wb.close()

    if ETH_CSV.exists():
        all_rows.extend(load_transaction_csv(ETH_CSV, "Ethereum Mainnet"))
    if BASE_CSV.exists():
        all_rows.extend(load_transaction_csv(BASE_CSV, "Base"))

    with open(OUTPUT_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        w.writerows(all_rows)

    print(f"Wrote {len(all_rows)} rows to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
