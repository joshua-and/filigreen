#!/usr/bin/env python3
"""Read an Excel (.xlsx) file and print its contents to stdout."""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install dependencies: pip install -r requirements.txt", file=sys.stderr)
    sys.exit(1)


def main():
    repo_root = Path(__file__).resolve().parent.parent
    path = repo_root / "raw" / "bank_for_daos_capital_tracking.xlsx"
    if not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== Sheet: {sheet_name} ===\n")
        for row in ws.iter_rows(values_only=True):
            print("\t".join(str(c) if c is not None else "" for c in row))
    wb.close()


if __name__ == "__main__":
    main()
